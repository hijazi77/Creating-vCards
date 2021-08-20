from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

name = [] 
phone = []
first = []
last = []

wb = load_workbook('contacts.xlsx')
ws = wb.active
input = int(input('How many contact you need in each file: '))
c = 1
for i in range(ws.max_row):
  name.append(ws[f'A{c}'].value)
  first.append(ws[f'B{c}'].value)
  last.append(ws[f'C{c}'].value)
  phone.append(ws[f'D{c}'].value)
  c+=1
  

file = 1

for i in range(int(ws.max_row/input)):
  with open(f'folder/file number-{file}.vcf', 'w' , encoding="utf-8") as myfile:
    count = 0
    for contact in range(input):
      myfile.write( 'BEGIN:VCARD' + "\n")
      myfile.write( 'VERSION:2.1' + "\n")
      myfile.write( 'N:' + str(last[count]) + ';' + (first[count]) + "\n")
      myfile.write( 'FN:' +str(name[count])+ "\n")
      myfile.write( 'ORG:' + 'hijazi77' + "\n")
      myfile.write( 'TEL;CELL:' + str(phone[count])+ "\n")
      myfile.write( 'END:VCARD' + "\n")
      myfile.write( "\n")
      name.pop(count)
      first.pop(count)
      last.pop(count)
      phone.pop(count)
      count+= 1
    myfile.close()
  file +=1

  print(f'file number -{file}- has been created')